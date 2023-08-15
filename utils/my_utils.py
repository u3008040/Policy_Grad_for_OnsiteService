# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
import os


def dijkstra(G, start, end=None):
    from pqdict import PQDict
    # G is the graph, start is the start location, end is the destination.
    # output is path. Path is in a form of list, each number is a node. it is impossible to derive the distance between
    # two intermediary points.
    if start == end:
        return 0, [start, end]
    start = int(start)
    try:
        end = int(end)
    except:
        pass
    inf = float('inf')
    D = {start: 0}  # mapping of nodes to their dist from start
    Q = PQDict(D)  # priority queue for tracking min shortest path
    P = {}  # mapping of nodes to their direct predecessors
    U = set(G.keys())  # unexplored nodes
    while U:  # nodes yet to explore
        (v, d) = Q.popitem()  # node w/ min dist d on frontier
        D[v] = d  # est dijkstra greedy score
        U.remove(v)  # remove from unexplored
        if v == end: break
        # now consider the edges from v with an unexplored head -
        # we may need to update the dist of unexplored successors
        for w in G[v]:  # successors to v
            if w in U:  # then w is a frontier node
                d = D[v] + G[v][w]  # dgs: dist of start -> v -> w
                if d < Q.get(w, inf):
                    Q[w] = d  # set/update dgs
                    P[w] = v  # set/update predecessor
    finish = False
    nextnode = end
    path = [end]
    while finish != True:
        nextnode = P[nextnode]
        path.append(nextnode)
        if nextnode == start:
            finish = True
    return D[end], path[::-1]  # distance and path


def dir_check(path):
    """
    check weather the dir of the given path exists, if not, then create it
    """
    import os
    dir = path if os.path.isdir(path) else os.path.split(path)[0]
    if not os.path.exists(dir): os.makedirs(dir)

def get_empty_lst(lst_len):
    "get a list with lst_len, with each item a empty lst"
    return [[] for _ in range(lst_len)]

def create_workbook(path):
    "create a .xlsx workbook"
    import xlsxwriter
    dir_check(path)
    workbook = xlsxwriter.Workbook(path)
    #worksheet = workbook.add_worksheet()
    workbook.close()

def discount_rewards(rewards, gamma):
    r = np.array([gamma ** i * rewards[i]
                  for i in range(len(rewards))])
    # Reverse the array direction for cumsum and then
    # revert back to the original order
    r = r[::-1].cumsum()[::-1]
    return r - r.mean()

def flatten(t):
    return [item for sublist in t for item in sublist]


def write_xlsx(df: pd.DataFrame, fout: str, sheet_name: str,  server_version: bool):
    from openpyxl import load_workbook
    if not os.path.exists(fout): create_workbook(fout)
    book = load_workbook(fout)
    writer = pd.ExcelWriter(fout,  engine='openpyxl', mode='a',if_sheet_exists='overlay')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    start_row = 0
    if sheet_name in writer.sheets.keys(): start_row = writer.sheets[sheet_name].max_row
    df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
    writer.save()

#-*- coding: utf-8 -*-
class GraphKruskal:

    def __init__(self, vertices):
        self.V = vertices  # Кількість вершин
        self.graph = []  # Словник для значень графу

        self.assignments = 0  # Лічильник для кількості присвоювань
        self.comparisons = 0  # Лічильник для кількості порівнянь

    # Функція для додавання ребра графу
    def addEdge(self, u, v, w):
        self.graph.append([u, v, w])

    # Рекурсивна функція, яка викликається щоб попередити зациклювання
    # (використовує метод стискання шляху)
    def find(self, parent, i):
        if parent[i] == i:
            self.comparisons += 1
            return i
        return self.find(parent, parent[i])

    # Функція, яка об'єднує корені х та у за рангом
    def union(self, parent, rank, x, y):
        xroot = self.find(parent, x)
        yroot = self.find(parent, y)
        self.assignments += 2

        # Додаємо дерево меншого рангу під корінь дерева більшого рангу
        if rank[xroot] < rank[yroot]:
            parent[xroot] = yroot
            self.comparisons += 1
            self.assignments += 1

        elif rank[xroot] > rank[yroot]:
            parent[yroot] = xroot
            self.comparisons += 1
            self.assignments += 1

        # Якщо ранги однакові, вершини відносяться до одного кореня, збільшуємо ранг на 1
        else:
            parent[yroot] = xroot
            rank[xroot] += 1
            self.comparisons += 1
            self.assignments += 1

    # Головна функція, що будує мінімальне остовне дерево за Крускалом і повпртає його вагу
    def Kruskal(self):
        result = []  # Тут буде зберігатися результат мінімального дерева

        # Індекс для відсортованих ребер
        i = 0

        # Індекс для result[]
        e = 0

        parent = []
        rank = []

        # Крок 1: сортуємо граф за зростанням
        self.graph = sorted(self.graph,
                            key=lambda item: item[2])

        # Створюємо набір вершин
        for node in range(self.V):
            parent.append(node)
            self.assignments += 1
            rank.append(0)

        # Алгоритм завершує роботу коли кількість вершин підграфа співпадає з кількістю вершин початкового графа
        while e < self.V - 1:
            self.comparisons += 1

            # Крок 2: беремо найменшу вершину з відсортованих і йдемо до кінця масиву
            u, v, w = self.graph[i]
            i = i + 1
            x = self.find(parent, u)
            y = self.find(parent, v)

            self.assignments += 3

            # Якщо додавання вершини не викликає зациклювання, то додаємо її до результату і продовжуємо
            if x != y:
                self.comparisons += 1
                e = e + 1
                result.append([u, v, w])
                self.union(parent, rank, x, y)
            # В іншому випадку вершину не додаємо

        minimumCost = 0
        string = []
        for u, v, weight in result:
            minimumCost += weight
            string.append([u ,v])# ,weight)
        return string, minimumCost, self.assignments, self.comparisons


def is_all_none(lst: list):
    cnt = 0
    for x in lst:
        if x is None: cnt+=1
    return cnt == len(lst)


def delete_by_idx(lst, delete_idx):
    "delete_ data of a list by given index"
    return [x for idx, x in enumerate(lst) if idx not in delete_idx]


from multiprocessing import Pool
def multi_thread_work(parameter_queue, function_name, thread_number=5):
    pool = Pool(thread_number)
    result = pool.map(function_name, parameter_queue)
    pool.close()
    pool.join()
    return result


# Set up lists to hold results
class Exp_Buffer():
    def __init__(self, buffer_size):
        self.total_rewards = []
        self.batch_rewards = []  # average reward along this batch
        self.batch_actions = []
        self.batch_states = []
        self.batch_graphs = []
        self.logprobs = [[] for _ in range(buffer_size)]
        self.batch_counter = 0
        self.state_counter = 0
        self.entropy = [[] for _ in range(buffer_size)]
        self.buffer_size = buffer_size

    def clear_logprobs(self):
        self.logprobs = [[] for _ in range(self.buffer_size)]
        self.entropy = [[] for _ in range(self.buffer_size)]

    def clear_rewards(self):
        self.batch_rewards = []
        self.total_rewards = []
        self.batch_counter = 0

    def add_new_instance(self, action, reward):
        self.total_rewards.append(reward)
        self.batch_actions.append(action)
        self.batch_counter += 1
        if self.batch_counter > self.buffer_size:
            del self.total_rewards[0]
            del self.batch_actions[0]
            self.batch_counter -= 1
        #print(len(self.total_rewards))

    def add_new_state(self, state, graph):
        self.batch_states = state
        self.batch_graphs = graph
        self.state_counter += 1
        if self.state_counter > 1:  # can only store one in the buffer
            self.state_counter -= 1

    def add_batch_rewards(self, averagereward, num):
        self.batch_rewards = self.batch_rewards + [averagereward for _ in range(num)]

    def get_instance(self):
        return self.batch_states, self.batch_graphs, self.total_rewards, self.batch_rewards, self.logprobs

import sys
class Logger(object):
    def __init__(self):
        self.terminal = sys.stdout  # stdout
        self.file = None

    def open(self, file, mode=None):
        if mode is None: mode = 'w'
        self.file = open(file, mode)

    def write(self, message, is_terminal=True, is_file=True):
        if '\r' in message: is_file = False

        if is_terminal:
            self.terminal.write(message)
            self.terminal.flush()
            # time.sleep(1)

        if is_file:
            self.file.write(message)
            self.file.flush()


